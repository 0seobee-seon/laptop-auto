"""
야근수당 자동 입력 스크립트 (범용)

1번 파일(야근현황 xlsx)에서 집계 데이터를 읽어
2번 파일(야근수당내역 xls)의 해당 월 시트에 자동으로 입력합니다.

■ 사용법
  python 야근수당_자동입력.py

■ 다른 부서에 적용하는 법
  아래 [사용자 설정] 섹션만 수정하세요.
  TARGET_FILE = None 으로 두면 새 xlsx 파일을 자동 생성합니다.
"""

import os
import re
import sys
import subprocess
import pandas as pd
import xlrd
from xlutils.copy import copy as xl_copy
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════
#  [사용자 설정] ─ 여기만 수정하면 됩니다
# ══════════════════════════════════════════════════

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 1번 파일: 야근현황 xlsx (출처 데이터)
SOURCE_FILE = os.path.join(BASE_DIR, "본사 설계 5월(1번).xlsx")

# 2번 파일: 야근수당내역 xls (입력 대상)
#   • 기존 파일 경로를 지정하면 해당 파일을 복사 후 업데이트
#   • None 으로 두면 새 xlsx 파일을 자동 생성 (서식 포함)
TARGET_FILE = os.path.join(BASE_DIR, "설계본부야근수당내역(2번).xls")

# 출력 파일 경로
OUTPUT_FILE = os.path.join(BASE_DIR, "설계본부야근수당내역_업데이트.xls")

# 2번 파일에서 입력할 시트명 (None이면 첫 번째 시트 사용)
TARGET_SHEET = "2026.05월"

# 1번 파일에서 건너뛸 시트 (데이터 시트가 아닌 것들)
SKIP_SHEETS = {"설정", "사용설명서"}

# 시트명 → 부서명 매핑
#   • 비워두면 시트명을 그대로 부서명으로 사용
#   • 여러 시트를 같은 부서로 묶을 수 있음
DEPT_MAP = {
    "설계1본부": "설계본부",
    "설계2본부": "설계본부",
    "구조팀":    "구조팀",
}

# 특정 직원을 다른 부서로 재분류
#   예) 설계1본부 시트에 있지만 실제로는 조경팀인 경우
#   비워두면 무시
SPECIAL_DEPT = {
    "박종민": "조경팀",
    "허경원": "조경팀",
}

# 고정 포괄임금 시간 (초과분만 추가 수당 지급)
FIXED_EXT  = 33   # 고정 연장·휴일 시간
FIXED_NIGT =  3   # 고정 야간 시간

# ── 누적 기록(시트 추가) 옵션 ────────────────────────
# TARGET_FILE 이 .xlsx 이고 TARGET_SHEET 가 파일에 없으면:
#   AUTO_CREATE_SHEET = True  → 기존 최신 월 시트를 복사해서 새 시트 생성
#   AUTO_CREATE_SHEET = False → 오류로 종료
AUTO_CREATE_SHEET = True

# TARGET_FILE 이 .xls 인 경우:
#   AUTO_CONVERT_XLS = True  → Excel 자동으로 .xlsx로 변환 후 시트 작업
#   AUTO_CONVERT_XLS = False → 기존 .xls 그대로 (시트 추가 불가, 기존 시트만 업데이트)
AUTO_CONVERT_XLS = True

# 새 시트 생성 시 템플릿으로 쓸 시트 (None이면 가장 최근 월 시트 자동 선택)
TEMPLATE_SHEET = None

# ══════════════════════════════════════════════════
#  내부 로직 (수정 불필요)
# ══════════════════════════════════════════════════

DECIMAL = 2   # 소수점 자리수


def r(v):
    """숫자 반올림 (전역 DECIMAL 자리)"""
    try:
        return round(float(v), DECIMAL)
    except Exception:
        return 0.0


def _find_summary_cols(df):
    """row 2 헤더를 스캔해서 집계 컬럼 위치를 자동으로 찾는다.

    파일마다 컬럼 수가 다를 수 있어 (72열, 74열 등) 동적으로 탐색.
    반환: {'day','night','hol_n','hol_e','comp'} → 컬럼 인덱스
    """
    cols = {}
    if df.shape[0] < 3:
        return cols
    for c in range(df.shape[1]):
        v = df.iloc[2, c]
        if pd.isna(v): continue
        h = str(v).replace('\n', '').replace(' ', '')
        if '평일야근' in h:
            cols['day'] = c
        elif '심야야근' in h:
            cols['night'] = c
        elif '휴일정상' in h:
            cols['hol_n'] = c
        elif '휴일초과' in h:
            cols['hol_e'] = c
        elif '보상휴가' in h:
            cols['comp'] = c
    return cols


def extract_overtime(source_path):
    """1번 파일에서 직원별 야근 집계 추출 (컬럼 자동 감지)"""
    xf = pd.ExcelFile(source_path)
    data_sheets = [s for s in xf.sheet_names if s not in SKIP_SHEETS]

    if not data_sheets:
        print("[오류] 1번 파일에서 데이터 시트를 찾지 못했습니다.")
        sys.exit(1)

    print(f"    감지된 데이터 시트: {data_sheets}")

    employees = {}

    for sheet in data_sheets:
        dept = DEPT_MAP.get(sheet, sheet)
        df = pd.read_excel(source_path, sheet_name=sheet, header=None)

        # 집계 컬럼 자동 감지
        sc = _find_summary_cols(df)
        required = {'day','night','hol_n','hol_e','comp'}
        if not required.issubset(sc):
            missing = required - set(sc)
            print(f"    [경고] '{sheet}' 시트에서 집계 컬럼 누락: {missing}")
            continue
        print(f"      [{sheet}] 집계컬럼: 평일={sc['day']} 심야={sc['night']} "
              f"휴일정상={sc['hol_n']} 휴일초과={sc['hol_e']} 보상={sc['comp']}")

        for i in range(6, len(df) - 1, 2):
            num   = df.iloc[i, 0]
            name  = df.iloc[i, 2]
            title = df.iloc[i, 1]

            if pd.isna(name) or pd.isna(num):
                continue

            name  = str(name).strip()
            title = "" if pd.isna(title) else str(title).strip()

            def val(row, col):
                x = df.iloc[row, col]
                return 0.0 if pd.isna(x) else float(x)

            # 소수 집계 행(i+1) - 동적 컬럼 인덱스 사용
            ext  = val(i+1, sc['day']) + val(i+1, sc['hol_n']) + val(i+1, sc['hol_e'])
            ngt  = val(i+1, sc['night'])
            comp = val(i,   sc['comp'])

            actual_dept = SPECIAL_DEPT.get(name, dept)

            employees[name] = {
                "dept":     actual_dept,
                "title":    title,
                "ext":      r(ext),
                "ngt":      r(ngt),
                "sub":      r(ext + ngt),
                "add_ext":  r(max(ext - FIXED_EXT,  0)),
                "add_ngt":  r(max(ngt - FIXED_NIGT, 0)),
                "comp":     r(comp),
            }

    return employees


# ─── 2번 파일 업데이트 (기존 xls 복사) ───────────────────────────

COL = {
    "no": 0, "dept": 1, "title": 2, "name": 3,
    "ext": 4, "ngt": 5, "sub": 6,
    "fext": 7, "fngt": 8,
    "add_ext": 9, "add_ngt": 10, "comp": 11, "proj": 12,
}


def update_existing(target_path, emp_data, sheet_name, output_path):
    """기존 xls 파일을 복사하여 데이터 업데이트 후 저장"""
    rb = xlrd.open_workbook(target_path, formatting_info=True)

    if sheet_name not in rb.sheet_names():
        print(f"[오류] 시트 '{sheet_name}' 없음. 사용 가능: {rb.sheet_names()}")
        sys.exit(1)

    wb = xl_copy(rb)
    sheet_idx = rb.sheet_names().index(sheet_name)
    ws = wb.get_sheet(sheet_idx)
    rs = rb.sheet_by_name(sheet_name)

    updated, skipped = [], []
    dept_totals = {}

    for row in range(3, rs.nrows):
        raw_name = str(rs.cell_value(row, COL["name"])).strip()
        raw_dept = str(rs.cell_value(row, COL["dept"])).strip()

        if not raw_name or raw_name in ("nan", "성명"):
            continue
        # 소계·합계 행 건너뜀 ("소     계", "합   계   시  간" 등)
        raw_no = str(rs.cell_value(row, COL["no"])).strip()
        if ("소" in raw_dept and "계" in raw_dept) or "합" in raw_no:
            continue

        emp = emp_data.get(raw_name) or emp_data.get(raw_name.replace(" ", ""))
        if emp is None:
            skipped.append(raw_name)
            continue

        ws.write(row, COL["ext"],     emp["ext"])
        ws.write(row, COL["ngt"],     emp["ngt"])
        ws.write(row, COL["sub"],     emp["sub"])
        ws.write(row, COL["add_ext"], emp["add_ext"])
        ws.write(row, COL["add_ngt"], emp["add_ngt"])
        ws.write(row, COL["comp"],    emp["comp"])

        d = emp["dept"]
        if d not in dept_totals:
            dept_totals[d] = {"add_ext": 0, "add_ngt": 0, "comp": 0}
        dept_totals[d]["add_ext"] += emp["add_ext"]
        dept_totals[d]["add_ngt"] += emp["add_ngt"]
        dept_totals[d]["comp"]    += emp["comp"]

        updated.append(raw_name)

    # 소계·합계 행 업데이트
    last_dept, all_ae, all_an, all_co = "", 0, 0, 0
    for row in range(3, rs.nrows):
        raw_dept = str(rs.cell_value(row, COL["dept"])).strip()
        raw_no   = str(rs.cell_value(row, COL["no"])).strip()

        is_subtotal = ("소" in raw_dept and "계" in raw_dept)
        is_total    = "합" in raw_no

        if is_subtotal:
            t = dept_totals.get(last_dept, {})
            ae = r(t.get("add_ext", 0))
            an = r(t.get("add_ngt", 0))
            co = r(t.get("comp", 0))
            ws.write(row, COL["add_ext"], ae)
            ws.write(row, COL["add_ngt"], an)
            ws.write(row, COL["comp"],    co)
            all_ae += ae; all_an += an; all_co += co
        elif is_total:
            ws.write(row, COL["add_ext"], r(all_ae))
            ws.write(row, COL["add_ngt"], r(all_an))
            ws.write(row, COL["comp"],    r(all_co))
        else:
            d = str(rs.cell_value(row, COL["dept"])).strip()
            if d and d not in ("nan",):
                last_dept = d

    wb.save(output_path)
    return updated, skipped


# ─── xls → xlsx 자동 변환 (Excel COM 사용) ───────────────────────

def _find_excel_exe():
    """Excel 실행 파일 찾기"""
    candidates = [
        r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office16\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office15\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office15\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office14\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office14\EXCEL.EXE",
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def convert_xls_to_xlsx(xls_path, xlsx_path=None):
    """Microsoft Excel을 통해 .xls 를 .xlsx 로 변환한다.

    이미 최신 .xlsx 가 존재하면 변환하지 않고 그대로 반환.
    """
    out_dir = os.path.dirname(xls_path) or "."
    base = os.path.splitext(os.path.basename(xls_path))[0]
    xlsx_path = xlsx_path or os.path.join(out_dir, base + ".xlsx")

    # 캐시 - .xlsx가 .xls보다 새것이면 그대로 사용
    if os.path.exists(xlsx_path) and \
       os.path.getmtime(xlsx_path) >= os.path.getmtime(xls_path):
        return xlsx_path

    try:
        import win32com.client as wcom
    except ImportError:
        raise RuntimeError(
            "pywin32 미설치 → 'pip install pywin32' 후 다시 실행하거나, "
            "수동으로 Excel에서 .xls → .xlsx 다른 이름 저장 후 TARGET_FILE 경로 갱신")

    excel = wcom.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(xls_path))
        # 51 = xlOpenXMLWorkbook (.xlsx)
        wb.SaveAs(os.path.abspath(xlsx_path), FileFormat=51)
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()

    if not os.path.exists(xlsx_path):
        raise RuntimeError(f"변환 실패: {xlsx_path}")
    return xlsx_path


# ─── xlsx 파일에 시트 업데이트 / 새 시트 추가 ──────────────────

def _make_period_label(sheet_name):
    """'2026.06월' → '2026-05-21 ~ 2026-06-20'
    (원본 1번 파일의 책정기간 표기 형식)
    """
    m = re.search(r"(\d{4}).*?(\d{1,2})", sheet_name)
    if not m:
        return sheet_name
    y = int(m.group(1)); mo = int(m.group(2))
    py, pm = (y, mo - 1) if mo > 1 else (y - 1, 12)
    return f"{py}-{pm:02d}-21 ~ {y}-{mo:02d}-20"


def _pick_template_sheet(wb, preferred=None):
    """템플릿 시트 선택. preferred 우선, 없으면 가장 최근 월 시트."""
    if preferred and preferred in wb.sheetnames:
        return preferred
    # 월 형식 시트 자동 선택
    month_re = re.compile(r"(\d{4}).*?(\d{1,2})")
    candidates = []
    for s in wb.sheetnames:
        m = month_re.search(s)
        if m:
            candidates.append((int(m.group(1)), int(m.group(2)), s))
    if candidates:
        candidates.sort()
        return candidates[-1][2]   # 가장 최근
    return wb.sheetnames[0]


def update_xlsx_workbook(target_path, emp_data, sheet_name, output_path=None,
                         template_sheet=None, auto_create=True,
                         move_to_index=None, period_label=None):
    """openpyxl로 xlsx 파일의 시트를 업데이트하거나 새로 생성한다.

    move_to_index: 새 시트 생성 시 워크북 내 위치(0=맨 앞). None이면 끝에 추가.
    period_label : 책정기간 라벨(E2 셀에 기록). None이면 sheet_name에서 추출.
    """
    wb = openpyxl.load_workbook(target_path)
    newly_created = False

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"    기존 시트 '{sheet_name}' 업데이트")
    elif auto_create:
        tpl = _pick_template_sheet(wb, template_sheet)
        ws = wb.copy_worksheet(wb[tpl])
        ws.title = sheet_name
        newly_created = True

        # 기간 라벨 갱신 (E2, 병합된 셀의 좌상단)
        label = period_label if period_label else _make_period_label(sheet_name)
        ws.cell(row=2, column=5, value=label)

        # 데이터 컬럼 초기화 (E, F, G, J, K, L)
        for row in range(4, ws.max_row + 1):
            for col in (5, 6, 7, 10, 11, 12):
                c = ws.cell(row=row, column=col)
                if isinstance(c.value, (int, float)):
                    c.value = 0
        print(f"    새 시트 '{sheet_name}' 생성 (템플릿: '{tpl}')")
    else:
        raise ValueError(f"시트 '{sheet_name}' 없음 / AUTO_CREATE_SHEET=False")

    # 새 시트를 지정한 위치로 이동
    if newly_created and move_to_index is not None:
        current_idx = wb.sheetnames.index(sheet_name)
        target_idx  = max(0, min(move_to_index, len(wb.sheetnames) - 1))
        offset = target_idx - current_idx
        if offset != 0:
            wb.move_sheet(ws, offset=offset)
            print(f"    시트 위치 이동: {current_idx} → {target_idx}")

    # 직원 매칭하여 데이터 채우기
    updated, skipped = [], []
    dept_totals = {}

    for row in range(4, ws.max_row + 1):
        name = str(ws.cell(row=row, column=4).value or "").strip()
        dept = str(ws.cell(row=row, column=2).value or "").strip()
        no   = str(ws.cell(row=row, column=1).value or "").strip()

        if not name or name in ("성명", "nan"):
            continue
        if ("소" in dept and "계" in dept) or "합" in no:
            continue

        emp = emp_data.get(name) or emp_data.get(name.replace(" ", ""))
        if emp is None:
            skipped.append(name)
            continue

        ws.cell(row=row, column=5).value  = emp["ext"]
        ws.cell(row=row, column=6).value  = emp["ngt"]
        ws.cell(row=row, column=7).value  = emp["sub"]
        ws.cell(row=row, column=10).value = emp["add_ext"]
        ws.cell(row=row, column=11).value = emp["add_ngt"]
        ws.cell(row=row, column=12).value = emp["comp"]

        d = emp["dept"]
        if d not in dept_totals:
            dept_totals[d] = {"add_ext": 0, "add_ngt": 0, "comp": 0}
        dept_totals[d]["add_ext"] += emp["add_ext"]
        dept_totals[d]["add_ngt"] += emp["add_ngt"]
        dept_totals[d]["comp"]    += emp["comp"]
        updated.append(name)

    # 소계·합계 행 갱신
    last_dept = ""
    overall = {"add_ext": 0, "add_ngt": 0, "comp": 0}
    for row in range(4, ws.max_row + 1):
        dept = str(ws.cell(row=row, column=2).value or "").strip()
        no   = str(ws.cell(row=row, column=1).value or "").strip()

        if "소" in dept and "계" in dept:
            t = dept_totals.get(last_dept, {})
            ae = r(t.get("add_ext", 0)); an = r(t.get("add_ngt", 0)); co = r(t.get("comp", 0))
            ws.cell(row=row, column=10).value = ae
            ws.cell(row=row, column=11).value = an
            ws.cell(row=row, column=12).value = co
            overall["add_ext"] += ae; overall["add_ngt"] += an; overall["comp"] += co
        elif "합" in no:
            ws.cell(row=row, column=10).value = r(overall["add_ext"])
            ws.cell(row=row, column=11).value = r(overall["add_ngt"])
            ws.cell(row=row, column=12).value = r(overall["comp"])
        elif dept and dept != "nan":
            last_dept = dept

    out = output_path or target_path
    wb.save(out)
    return updated, skipped


# ─── 새 파일 생성 (target None인 경우) ──────────────────────────
# 원본 2번 파일 서식 기준
#   배경색  : FFCC99 (오렌지) - 헤더/소계/합계
#             FFCC00 (노랑)   - 추가연장·추가야간·보상휴가 열
#   폰트    : 굴림
#   테두리  : medium (2) / thin (1) - xlrd 기준
#   열 너비 : xlrd computed_column_width 값 / 256
#   행 높이 : xlrd 단위 / 20 = pt

_ORANGE  = "FFCC99"
_YELLOW  = "FFCC00"
_FONT    = "굴림"
_SIZE    = 12   # 모든 글자 크기 통일

# 열 너비 (사용자 수정 값 반영)
#   A    B     C      D       E      F     G     H       I      J      K      L     M
_COL_W = [5.0, 10.0, 7.125, 10.875, 20.75, 13.0, 13.0, 11.125, 10.25, 15.75, 10.25, 15.0, 17.0]

# 행 높이 (사용자 수정 값 반영)
_RH = {"unit": 15.95, "h1": 36.0, "h2": 20.1, "emp": 21.95, "sub": 24.0, "total": 27.95}


def _f(name=_FONT, size=None, bold=False):
    return Font(name=name, size=size if size is not None else _SIZE, bold=bold)


def _a(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _b(t=0, b=0, l=0, r=0):
    _s = {0: None, 1: "thin", 2: "medium"}
    return Border(
        top=Side(style=_s[t]) if t else Side(style=None),
        bottom=Side(style=_s[b]) if b else Side(style=None),
        left=Side(style=_s[l]) if l else Side(style=None),
        right=Side(style=_s[r]) if r else Side(style=None),
    )


def _fl(color=None):
    return PatternFill("solid", fgColor=color) if color else PatternFill(fill_type=None)


def _wc(ws, row, col, val=None, fnt=None, fl=None, bd=None, al=None):
    c = ws.cell(row=row, column=col, value=val)
    if fnt: c.font = fnt
    if fl is not None: c.fill = fl
    if bd: c.border = bd
    if al: c.alignment = al
    return c


def create_new(emp_data, sheet_name, output_path, period_str=""):
    """직원 명단과 집계 데이터로 새 xlsx 파일 생성 (2번 파일 서식 동일)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── 열 너비 ─────────────────────────────────────────
    for i, w in enumerate(_COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    period_label = period_str or _make_period_label(sheet_name)

    # ── Row 1: (단위:원) ──────────────────────────────
    ws.row_dimensions[1].height = _RH["unit"]
    for col in range(1, 14):
        _wc(ws, 1, col, fnt=_f(), fl=_fl(), bd=_b(), al=_a())
    _wc(ws, 1, 13, "(단위:원)", fnt=_f(), fl=_fl(), al=_a("right"))

    # ── Row 2: 주 헤더 ────────────────────────────────
    ws.row_dimensions[2].height = _RH["h1"]
    # (col, value, bold, bg, t,b,l,r, wrap, ha)
    h1 = [
        (1,  "구분",              True, _ORANGE, 2,0,2,1, False, "center"),
        (2,  "부서",              True, _ORANGE, 2,0,1,1, False, "center"),
        (3,  "직책",              True, _ORANGE, 2,0,1,1, False, "center"),
        (4,  "성명",              True, _ORANGE, 2,0,1,1, False, "center"),
        (5,  period_label,         True, _ORANGE, 2,1,1,0, False, "center"),
        (6,  None,                 True, _ORANGE, 2,1,0,0, False, "center"),
        (7,  None,                 True, _ORANGE, 2,1,0,1, False, "center"),
        (8,  "고정연장.\n휴일시간", True, _ORANGE, 2,0,1,1, True,  "center"),
        (9,  "고정야간\n시간",      True, _ORANGE, 2,0,1,1, True,  "center"),
        (10, "추가연장.휴일\n근무시간", True, _ORANGE, 2,0,1,1, True, "center"),
        (11, "추가야간\n근무시간",  True, _ORANGE, 2,0,1,1, True,  "center"),
        (12, "보상휴가시간",        True, _ORANGE, 2,0,1,1, False, "center"),
        (13, "프로젝트명",         True, _ORANGE, 2,0,1,1, False, "center"),
    ]
    for col,val,bold,bg,t,b,l,ri,wrap,ha in h1:
        _wc(ws,2,col,val,_f(bold=bold),_fl(bg),_b(t,b,l,ri),_a(ha,"center",wrap))

    # ── Row 3: 서브 헤더 ─────────────────────────────
    ws.row_dimensions[3].height = _RH["h2"]
    h2 = [
        (1,  None,           True, _ORANGE, 0,2,2,1, False),
        (2,  None,           True, _ORANGE, 0,2,1,1, False),
        (3,  None,           True, _ORANGE, 0,2,1,1, False),
        (4,  None,           True, _ORANGE, 0,2,1,1, False),
        (5,  "연장.휴일근무",  True, _ORANGE, 1,2,1,1, False),
        (6,  "야간근무",      True, _ORANGE, 1,2,1,1, False),
        (7,  "소 계",        True, _ORANGE, 1,2,1,1, False),
        (8,  None,            True, _ORANGE, 0,2,1,1, True),
        (9,  None,            True, _ORANGE, 0,2,1,1, True),
        (10, None,            True, _ORANGE, 0,2,1,1, True),
        (11, None,            True, _ORANGE, 0,2,1,1, True),
        (12, None,            True, _ORANGE, 0,2,1,1, False),
        (13, None,           True, _ORANGE, 0,2,1,1, False),
    ]
    for col,val,bold,bg,t,b,l,ri,wrap in h2:
        _wc(ws,3,col,val,_f(bold=bold),_fl(bg),_b(t,b,l,ri),_a("center","center",wrap))

    # 헤더 병합
    for rng in ("A2:A3","B2:B3","C2:C3","D2:D3",
                "E2:G2",
                "H2:H3","I2:I3","J2:J3","K2:K3","L2:L3","M2:M3"):
        ws.merge_cells(rng)

    # ── 직원 데이터 행 ────────────────────────────────
    dept_order, dept_emps = [], {}
    for name, info in emp_data.items():
        d = info["dept"]
        if d not in dept_emps:
            dept_order.append(d)
            dept_emps[d] = []
        dept_emps[d].append((name, info))

    cur = 4
    overall = {"add_ext": 0, "add_ngt": 0, "comp": 0}

    for dept in dept_order:
        emps = dept_emps[dept]
        dtot = {"add_ext": 0, "add_ngt": 0, "comp": 0}

        for seq, (name, info) in enumerate(emps, 1):
            ws.row_dimensions[cur].height = _RH["emp"]
            # (col, val, bg, t,b,l,r)
            emp_row = [
                (1,  seq,              None,    1,1,2,1),
                (2,  dept,             None,    1,1,0,1),
                (3,  info["title"],    None,    1,1,1,1),
                (4,  name,             None,    1,1,1,1),
                (5,  info["ext"],      None,    1,1,1,1),
                (6,  info["ngt"],      None,    1,1,1,1),
                (7,  info["sub"],      None,    1,1,1,1),
                (8,  FIXED_EXT,        None,    1,1,1,1),
                (9,  FIXED_NIGT,       None,    1,1,1,1),
                (10, info["add_ext"], _YELLOW,  1,1,1,1),
                (11, info["add_ngt"], _YELLOW,  1,1,1,1),
                (12, info["comp"],    _YELLOW,  1,1,1,1),
                (13, "-",              None,    1,1,1,1),
            ]
            for col,val,bg,t,b,l,ri in emp_row:
                _wc(ws,cur,col,val,_f(),_fl(bg),_b(t,b,l,ri),_a())

            dtot["add_ext"] += info["add_ext"]
            dtot["add_ngt"] += info["add_ngt"]
            dtot["comp"]    += info["comp"]
            cur += 1

        # ── 소계 행 ───────────────────────────────────
        ws.row_dimensions[cur].height = _RH["sub"]
        _wc(ws,cur,1, None,              _f(bold=True), _fl(None),    _b(1,1,2,1), _a())
        _wc(ws,cur,2, "소     계",       _f(bold=True), _fl(_ORANGE), _b(1,1,1,0), _a())
        for col in range(3,9):
            _wc(ws,cur,col, None,         _f(bold=True), _fl(_ORANGE), _b(1,1,0,0), _a())
        _wc(ws,cur,9,  None,              _f(bold=True), _fl(_ORANGE), _b(1,1,0,1), _a())
        _wc(ws,cur,10, r(dtot["add_ext"]),_f(bold=True), _fl(_ORANGE), _b(1,1,1,1), _a())
        _wc(ws,cur,11, r(dtot["add_ngt"]),_f(bold=True), _fl(_ORANGE), _b(1,1,1,1), _a())
        _wc(ws,cur,12, r(dtot["comp"]),   _f(bold=True), _fl(_ORANGE), _b(1,1,1,1), _a())
        _wc(ws,cur,13, None,              _f(bold=True), _fl(_ORANGE), _b(1,1,1,1), _a())
        ws.merge_cells(f"B{cur}:H{cur}")   # B-H 병합

        overall["add_ext"] += dtot["add_ext"]
        overall["add_ngt"] += dtot["add_ngt"]
        overall["comp"]    += dtot["comp"]
        cur += 1

    # ── 합계 행 ───────────────────────────────────────
    ws.row_dimensions[cur].height = _RH["total"]
    _wc(ws,cur,1, "합   계   시  간", _f(bold=True), _fl(_ORANGE), _b(2,2,2,0), _a())
    for col in range(2, 9):
        _wc(ws,cur,col, None,            _f(bold=True), _fl(_ORANGE), _b(2,2,0,0), _a())
    _wc(ws,cur,9,  None,                 _f(bold=True), _fl(_ORANGE), _b(2,2,0,1), _a())
    _wc(ws,cur,10, r(overall["add_ext"]),_f(bold=True), _fl(_ORANGE), _b(2,2,1,1), _a())
    _wc(ws,cur,11, r(overall["add_ngt"]),_f(bold=True), _fl(_ORANGE), _b(2,2,1,1), _a())
    _wc(ws,cur,12, r(overall["comp"]),   _f(bold=True), _fl(_ORANGE), _b(2,2,1,1), _a())
    _wc(ws,cur,13, None,                 _f(),          _fl(_ORANGE), _b(2,2,1,1), _a())
    ws.merge_cells(f"A{cur}:I{cur}")   # A-I 병합

    # ── A4 세로 인쇄 설정 ────────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 여백 (inch 단위) ─ 위/아래 1cm, 좌/우 0.75cm
    from openpyxl.worksheet.page import PageMargins
    ws.page_margins = PageMargins(
        left=0.295, right=0.295,
        top=0.3937, bottom=0.3937,
        header=0.2, footer=0.2,
    )
    ws.print_options.horizontalCentered = True

    out = output_path if output_path.endswith(".xlsx") else output_path.replace(".xls", ".xlsx")
    wb.save(out)
    return out


# ─── 빈 양식 (공란 템플릿) 생성 ───────────────────────

def create_empty_template(sheet_name, output_path, num_rows=30, period_str=""):
    """데이터가 없는 빈 양식 파일을 생성한다.

    num_rows: 미리 만들 빈 직원 행 개수 (기본 30)
    """
    # create_new 와 동일한 헤더/스타일을 사용하기 위해
    # 더미 직원 데이터를 만들어서 create_new를 호출한 뒤,
    # 데이터 셀 값만 비우는 방식 대신 — 직접 작성한다.

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 열 너비
    for i, w in enumerate(_COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    period_label = period_str or _make_period_label(sheet_name)

    # Row 1: (단위:원)
    ws.row_dimensions[1].height = _RH["unit"]
    for col in range(1, 14):
        _wc(ws, 1, col, fnt=_f(), fl=_fl(), bd=_b(), al=_a())
    _wc(ws, 1, 13, "(단위:원)", fnt=_f(), fl=_fl(), al=_a("right"))

    # Row 2: 주 헤더
    ws.row_dimensions[2].height = _RH["h1"]
    h1 = [
        (1,  "구분",              True, _ORANGE, 2,0,2,1, False),
        (2,  "부서",              True, _ORANGE, 2,0,1,1, False),
        (3,  "직책",              True, _ORANGE, 2,0,1,1, False),
        (4,  "성명",              True, _ORANGE, 2,0,1,1, False),
        (5,  period_label,         True, _ORANGE, 2,1,1,0, False),
        (6,  None,                 True, _ORANGE, 2,1,0,0, False),
        (7,  None,                 True, _ORANGE, 2,1,0,1, False),
        (8,  "고정연장.\n휴일시간", True, _ORANGE, 2,0,1,1, True),
        (9,  "고정야간\n시간",      True, _ORANGE, 2,0,1,1, True),
        (10, "추가연장.휴일\n근무시간", True, _ORANGE, 2,0,1,1, True),
        (11, "추가야간\n근무시간",  True, _ORANGE, 2,0,1,1, True),
        (12, "보상휴가시간",        True, _ORANGE, 2,0,1,1, False),
        (13, "프로젝트명",         True, _ORANGE, 2,0,1,1, False),
    ]
    for col,val,bold,bg,t,b,l,ri,wrap in h1:
        _wc(ws,2,col,val,_f(bold=bold),_fl(bg),_b(t,b,l,ri),_a("center","center",wrap))

    # Row 3: 서브 헤더
    ws.row_dimensions[3].height = _RH["h2"]
    h2 = [
        (1,  None,           True, _ORANGE, 0,2,2,1, False),
        (2,  None,           True, _ORANGE, 0,2,1,1, False),
        (3,  None,           True, _ORANGE, 0,2,1,1, False),
        (4,  None,           True, _ORANGE, 0,2,1,1, False),
        (5,  "연장.휴일근무",  True, _ORANGE, 1,2,1,1, False),
        (6,  "야간근무",      True, _ORANGE, 1,2,1,1, False),
        (7,  "소 계",        True, _ORANGE, 1,2,1,1, False),
        (8,  None,            True, _ORANGE, 0,2,1,1, True),
        (9,  None,            True, _ORANGE, 0,2,1,1, True),
        (10, None,            True, _ORANGE, 0,2,1,1, True),
        (11, None,            True, _ORANGE, 0,2,1,1, True),
        (12, None,            True, _ORANGE, 0,2,1,1, False),
        (13, None,           True, _ORANGE, 0,2,1,1, False),
    ]
    for col,val,bold,bg,t,b,l,ri,wrap in h2:
        _wc(ws,3,col,val,_f(bold=bold),_fl(bg),_b(t,b,l,ri),_a("center","center",wrap))

    # 헤더 병합
    for rng in ("A2:A3","B2:B3","C2:C3","D2:D3",
                "E2:G2",
                "H2:H3","I2:I3","J2:J3","K2:K3","L2:L3","M2:M3"):
        ws.merge_cells(rng)

    # ── 빈 직원 행 ───────────────────────────────
    start_row = 4
    for i in range(num_rows):
        cur = start_row + i
        ws.row_dimensions[cur].height = _RH["emp"]
        # 구분(1), 부서, 직책, 성명, 연장, 야간, 소계, 고정연장(=33), 고정야간(=3), 추가연장, 추가야간, 보상휴가, 프로젝트
        empty_row = [
            (1,  i+1,        None,    1,1,2,1),
            (2,  None,       None,    1,1,0,1),
            (3,  None,       None,    1,1,1,1),
            (4,  None,       None,    1,1,1,1),
            (5,  None,       None,    1,1,1,1),
            (6,  None,       None,    1,1,1,1),
            (7,  None,       None,    1,1,1,1),
            (8,  FIXED_EXT,  None,    1,1,1,1),
            (9,  FIXED_NIGT, None,    1,1,1,1),
            (10, None,       _YELLOW, 1,1,1,1),
            (11, None,       _YELLOW, 1,1,1,1),
            (12, None,       _YELLOW, 1,1,1,1),
            (13, None,       None,    1,1,1,1),
        ]
        for col,val,bg,t,b,l,ri in empty_row:
            _wc(ws,cur,col,val,_f(),_fl(bg),_b(t,b,l,ri),_a())

    # ── 소계 행 (빈 양식이므로 합계 수식 미포함) ──────
    cur = start_row + num_rows
    ws.row_dimensions[cur].height = _RH["sub"]
    _wc(ws,cur,1, None,           _f(bold=True), _fl(None),    _b(1,1,2,1), _a())
    _wc(ws,cur,2, "소     계",    _f(bold=True), _fl(_ORANGE), _b(1,1,1,0), _a())
    for col in range(3,9):
        _wc(ws,cur,col, None,      _f(bold=True), _fl(_ORANGE), _b(1,1,0,0), _a())
    _wc(ws,cur,9,  None,           _f(bold=True), _fl(_ORANGE), _b(1,1,0,1), _a())
    # 소계 수식
    _wc(ws,cur,10, f"=SUM(J{start_row}:J{cur-1})", _f(bold=True), _fl(_ORANGE), _b(1,1,1,1), _a())
    _wc(ws,cur,11, f"=SUM(K{start_row}:K{cur-1})", _f(bold=True), _fl(_ORANGE), _b(1,1,1,1), _a())
    _wc(ws,cur,12, f"=SUM(L{start_row}:L{cur-1})", _f(bold=True), _fl(_ORANGE), _b(1,1,1,1), _a())
    _wc(ws,cur,13, None,           _f(bold=True), _fl(_ORANGE), _b(1,1,1,1), _a())
    ws.merge_cells(f"B{cur}:H{cur}")

    # ── 합계 행 ─────────────────────────────────
    sub_row = cur
    cur += 1
    ws.row_dimensions[cur].height = _RH["total"]
    _wc(ws,cur,1, "합   계   시  간", _f(bold=True), _fl(_ORANGE), _b(2,2,2,0), _a())
    for col in range(2, 9):
        _wc(ws,cur,col, None,            _f(bold=True), _fl(_ORANGE), _b(2,2,0,0), _a())
    _wc(ws,cur,9,  None,                 _f(bold=True), _fl(_ORANGE), _b(2,2,0,1), _a())
    # 합계는 소계 셀 참조
    _wc(ws,cur,10, f"=J{sub_row}", _f(bold=True), _fl(_ORANGE), _b(2,2,1,1), _a())
    _wc(ws,cur,11, f"=K{sub_row}", _f(bold=True), _fl(_ORANGE), _b(2,2,1,1), _a())
    _wc(ws,cur,12, f"=L{sub_row}", _f(bold=True), _fl(_ORANGE), _b(2,2,1,1), _a())
    _wc(ws,cur,13, None,           _f(),          _fl(_ORANGE), _b(2,2,1,1), _a())
    ws.merge_cells(f"A{cur}:I{cur}")

    # ── A4 세로 인쇄 설정 ──────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    from openpyxl.worksheet.page import PageMargins
    ws.page_margins = PageMargins(
        left=0.295, right=0.295,
        top=0.3937, bottom=0.3937,
        header=0.2, footer=0.2,
    )
    ws.print_options.horizontalCentered = True

    out = output_path if output_path.endswith(".xlsx") else output_path.replace(".xls", ".xlsx")
    wb.save(out)
    return out


# ══════════════════════════════════════════════════
#  실행
# ══════════════════════════════════════════════════

if __name__ == "__main__":
    print(f"[1] 1번 파일 읽는 중: {os.path.basename(SOURCE_FILE)}")
    emp_data = extract_overtime(SOURCE_FILE)
    print(f"    → 추출된 직원 수: {len(emp_data)}명")

    if TARGET_FILE and os.path.exists(TARGET_FILE):
        target = TARGET_FILE
        ext = os.path.splitext(target)[1].lower()

        # 1) .xls → .xlsx 자동 변환 (필요 시)
        if ext == ".xls" and AUTO_CONVERT_XLS:
            try:
                print(f"[2-a] xls → xlsx 자동 변환 중...")
                target = convert_xls_to_xlsx(target)
                print(f"      변환 완료: {os.path.basename(target)}")
                ext = ".xlsx"
            except Exception as e:
                print(f"      [경고] 자동 변환 실패: {e}")
                print(f"      → 기존 .xls 모드로 진행 (시트 추가 불가)")

        # 2) xlsx면 openpyxl로 업데이트 (시트 추가 가능)
        if ext == ".xlsx":
            print(f"[2-b] xlsx 시트 처리: {TARGET_SHEET}")
            updated, skipped = update_xlsx_workbook(
                target, emp_data, TARGET_SHEET,
                output_path=OUTPUT_FILE if OUTPUT_FILE != target else None,
                template_sheet=TEMPLATE_SHEET,
                auto_create=AUTO_CREATE_SHEET,
            )
            out_path = OUTPUT_FILE if OUTPUT_FILE != target else target
        else:
            # xls 모드 - 기존 시트만 업데이트 가능
            print(f"[2-b] xls 시트 업데이트: {TARGET_SHEET}")
            updated, skipped = update_existing(target, emp_data, TARGET_SHEET, OUTPUT_FILE)
            out_path = OUTPUT_FILE
    else:
        print(f"[2] 새 파일 생성 중...")
        out_path = create_new(emp_data, TARGET_SHEET, OUTPUT_FILE)
        updated = list(emp_data.keys())
        skipped = []

    print()
    print("=" * 55)
    print(f"  완료: {out_path}")
    print(f"  업데이트된 직원: {len(updated)}명")
    if skipped:
        print(f"  [주의] 1번 파일에 없는 직원 (0 유지): {skipped}")
    print("=" * 55)
