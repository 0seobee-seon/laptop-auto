#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
선엔지니어링 노트북 지급현황 마스터 파일 자동 최신화 스크립트

Usage:
    python update_laptop_master.py --hr <인사기록부.xlsx> [--new-equip <지급현황.xlsx>]
                                   [--master <마스터.xlsx>] [--out <출력.xlsx>]
"""
import argparse, json, re, sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"ERROR: 필수 라이브러리 없음 → {e}")
    print("실행: pip install pandas openpyxl")
    sys.exit(1)

DEFAULT_MASTER = (
    r"D:\D\김영섭\현장\감리현장 전산기기 납품내역서"
    r"\선엔지니어링_노트북지급현황_최신화.xlsx"
)

# ─────────────────────────────────────────────────────────────────────────────
# 유틸리티
# ─────────────────────────────────────────────────────────────────────────────

def norm(name: str) -> str:
    """이름 정규화: 공백·NBSP 제거"""
    return str(name).replace(" ", "").replace(" ", "").strip()

def cell_str(val) -> str:
    s = str(val).strip() if val is not None else ""
    return "" if s == "nan" else s

def excel_serial_to_date(val) -> str:
    """엑셀 시리얼 번호 → YYYY-MM-DD"""
    try:
        s = int(float(str(val)))
        if s > 0:
            return (datetime(1899, 12, 30) + timedelta(days=s)).strftime("%Y-%m-%d")
    except Exception:
        pass
    return ""

def parse_bdate(val) -> str:
    """생년월일 YYMMDD → YYYY-MM-DD"""
    s = str(val).strip().split(".")[0]
    if len(s) == 6:
        try:
            yy, mm, dd = int(s[:2]), s[2:4], s[4:]
            yyyy = 1900 + yy if yy >= 30 else 2000 + yy
            return f"{yyyy}-{mm}-{dd}"
        except Exception:
            pass
    return ""

def parse_date_key(s: str) -> str:
    """다양한 날짜 포맷 → 정렬용 YYYY-MM 키"""
    s = str(s).strip()
    patterns = [
        (r"(\d{4})년\s*(\d{1,2})월",  lambda m: f"{m.group(1)}-{int(m.group(2)):02d}"),
        (r"(\d{2})\.(\d{1,2})월",     lambda m: f"20{m.group(1)}-{int(m.group(2)):02d}"),
        (r"(\d{4})\.(\d{1,2})월",     lambda m: f"{m.group(1)}-{int(m.group(2)):02d}"),
        (r"(\d{4})-(\d{2})-(\d{2})",  lambda m: f"{m.group(1)}-{m.group(2)}"),
    ]
    for pat, fmt in patterns:
        m = re.match(pat, s)
        if m:
            try:
                return fmt(m)
            except Exception:
                pass
    return ""

# ─────────────────────────────────────────────────────────────────────────────
# 인사기록부 파싱
# ─────────────────────────────────────────────────────────────────────────────

def parse_hr_file(path: str):
    """
    Returns
    -------
    active : dict  {norm_name → {구분, 사번, 소속, 직종, 직책, 성명, 생년월일, 입사일}}
    resigned : set {norm_name}
    """
    xl = pd.ExcelFile(path)

    # (주)선종합 시트 탐색
    hr_sheet = next((s for s in xl.sheet_names if "선종합" in s), None)
    if hr_sheet is None:
        raise ValueError(
            f"'선종합' 포함 시트를 찾을 수 없습니다.\n"
            f"현재 시트 목록: {xl.sheet_names}"
        )

    active = {}
    df = pd.read_excel(path, sheet_name=hr_sheet, header=None)
    for i, row in df.iterrows():
        if i < 2:
            continue
        cols = [cell_str(row[c]) for c in range(min(19, len(row)))]
        name = norm(cols[5]) if len(cols) > 5 else ""
        if not name:
            continue
        active[name] = {
            "구분":    cols[0],
            "사번":    cols[1],
            "소속":    cols[2],
            "직종":    cols[3],
            "직책":    cols[4],
            "성명":    name,
            "생년월일": parse_bdate(cols[6]) if len(cols) > 6 else "",
            "입사일":  excel_serial_to_date(cols[8]) if len(cols) > 8 else "",
        }

    # 퇴사자 시트 탐색
    resigned = set()
    res_sheet = next((s for s in xl.sheet_names if "퇴사" in s), None)
    if res_sheet:
        df_res = pd.read_excel(path, sheet_name=res_sheet, header=None)
        for i, row in df_res.iterrows():
            if i < 1:
                continue
            n = norm(cell_str(row[5]) if len(row) > 5 else "")
            if n:
                resigned.add(n)

    return active, resigned

# ─────────────────────────────────────────────────────────────────────────────
# 전산기기 지급현황 파일 파싱
# ─────────────────────────────────────────────────────────────────────────────

def detect_format(df: pd.DataFrame):
    """
    포맷 자동 감지.
    Returns ('distribution'|'outgoing'|'edu'|'unknown', header_row_index)
    - distribution : 개인별 노트북 지급현황 양식 (성명·현장명·모델명·인수일자)
    - outgoing     : 선 노트북출고현황 양식 (모델명·S/N·날짜·성함·현장)
    - edu          : 교육용 지급 노트북 양식 (No.·제품명·시리얼넘버·비고·소속·이름)
    """
    for i in range(min(4, len(df))):
        row_vals = [str(v).strip() for v in df.iloc[i]]
        joined = " ".join(row_vals)
        # 포맷 C: 교육용 (시리얼넘버 또는 제품명+이름 조합)
        if "시리얼넘버" in joined or ("제품명" in joined and "이름" in joined):
            return "edu", i
        # 포맷 A: 개인별 지급현황 (성명·생년월일·현장명 키워드)
        if any(k in joined for k in ("생년월일", "현장명", "인수일자")):
            return "distribution", i
        if "성   명" in joined or "성명" in joined.replace(" ", ""):
            if "현장" in joined:
                return "distribution", i
        # 포맷 B: 출고현황 (성함·S/N 키워드)
        if "성함" in joined or "S/N" in joined or ("년/월" in joined and "성함" in joined):
            return "outgoing", i
    return "unknown", 0


def _date_from_filename(path: str) -> str:
    """파일명에서 날짜 추출 → 'YYYY년 MM월' 형식 반환. 실패 시 빈 문자열."""
    name = Path(path).stem  # 확장자 제외 파일명
    m = re.search(r"(\d{4})년\s*(\d{1,2})월", name)
    if m:
        return f"{m.group(1)}년{int(m.group(2)):02d}월"
    m = re.search(r"(\d{4})[-_\.](\d{1,2})", name)
    if m:
        return f"{m.group(1)}년{int(m.group(2)):02d}월"
    return ""


def parse_equip_file(path: str) -> list:
    """
    Returns
    -------
    list of {성명, 현장명, 모델명, 모델번호, 인수일자, 비고, date_key}
    """
    xl = pd.ExcelFile(path)
    assignments = []

    # 교육용 파일처럼 날짜가 파일명에 있을 때 사용
    file_date_str = _date_from_filename(path)
    file_date_key = parse_date_key(file_date_str)

    for sheet_name in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet_name, header=None)
        if df.empty:
            continue

        fmt, hrow = detect_format(df)

        if fmt == "distribution":
            # col 5=성명, 9=현장명, 10=모델명, 11=모델번호, 12=인수일자, 13=비고
            for i, row in df.iterrows():
                if i <= hrow:
                    continue
                cols = [cell_str(row[c]) for c in range(min(14, len(row)))]
                name = norm(cols[5]) if len(cols) > 5 else ""
                if not name:
                    continue
                date_str = cols[12] if len(cols) > 12 else ""
                assignments.append({
                    "성명":    name,
                    "현장명":  cols[9]  if len(cols) > 9  else "",
                    "모델명":  cols[10] if len(cols) > 10 else "",
                    "모델번호": cols[11] if len(cols) > 11 else "",
                    "인수일자": date_str,
                    "비고":    cols[13] if len(cols) > 13 else "",
                    "date_key": parse_date_key(date_str),
                })

        elif fmt == "outgoing":
            # col 0=모델명, 1=S/N, 2=날짜, 3=성함, 5=현장, 6=비고
            for i, row in df.iterrows():
                if i <= hrow:
                    continue
                cols = [cell_str(row[c]) for c in range(min(8, len(row)))]
                name = norm(cols[3]) if len(cols) > 3 else ""
                if not name or name in ("성함", ""):
                    continue
                date_str = cols[2] if len(cols) > 2 else ""
                assignments.append({
                    "성명":    name,
                    "현장명":  cols[5] if len(cols) > 5 else "",
                    "모델명":  cols[0] if len(cols) > 0 else "",
                    "모델번호": cols[1] if len(cols) > 1 else "",
                    "인수일자": date_str,
                    "비고":    cols[6] if len(cols) > 6 else "",
                    "date_key": parse_date_key(date_str),
                })

        elif fmt == "edu":
            # col 0=No., 1=제품명(모델명), 2=시리얼넘버, 3=비고, 4=소속(현장명), 5=이름(성명)
            # 날짜는 파일명에서 추출
            for i, row in df.iterrows():
                if i <= hrow:
                    continue
                cols = [cell_str(row[c]) for c in range(min(6, len(row)))]
                name = norm(cols[5]) if len(cols) > 5 else ""
                if not name or name in ("이름", ""):
                    continue
                note_parts = [p for p in [cols[3] if len(cols) > 3 else "", "교육용"] if p]
                assignments.append({
                    "성명":    name,
                    "현장명":  cols[4] if len(cols) > 4 else "",  # 소속을 현장명으로
                    "모델명":  cols[1] if len(cols) > 1 else "",
                    "모델번호": cols[2] if len(cols) > 2 else "",
                    "인수일자": file_date_str,
                    "비고":    " / ".join(note_parts),
                    "date_key": file_date_key,
                })
        # unknown 포맷은 건너뜀

    return assignments

# ─────────────────────────────────────────────────────────────────────────────
# 마스터 파일 로드
# ─────────────────────────────────────────────────────────────────────────────

def load_master(path: str) -> dict:
    """
    기존 마스터 파일에서 노트북 이력 로드.
    Returns {norm_name → {모델명, 모델번호, 현장명, 인수일자, 비고, in_sheet}}
    """
    master = {}
    p = Path(path)
    if not p.exists():
        return master

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  [경고] 마스터 파일 열기 실패: {e}")
        return master

    for sheet_name in ["노트북보유", "미지급"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(3, ws.max_row + 1):
            raw_name = ws.cell(row=r, column=6).value
            if not raw_name:
                continue
            name = norm(str(raw_name))
            master[name] = {
                "현장명":  str(ws.cell(row=r, column=10).value or ""),
                "모델명":  str(ws.cell(row=r, column=11).value or ""),
                "모델번호": str(ws.cell(row=r, column=12).value or ""),
                "인수일자": str(ws.cell(row=r, column=13).value or ""),
                "비고":    str(ws.cell(row=r, column=14).value or ""),
                "in_sheet": sheet_name,
            }
    return master

# ─────────────────────────────────────────────────────────────────────────────
# Excel 출력
# ─────────────────────────────────────────────────────────────────────────────

_HFILL = PatternFill("solid", start_color="4472C4", end_color="4472C4")
_HFONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=9)
_DFONT = Font(name="맑은 고딕", size=9)
_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN  = Side(style="thin", color="000000")
_BRD   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WIDTHS  = [5, 7, 12, 7, 8, 9, 11, 2, 12, 18, 12, 17, 10, 10]
_HEADERS = [
    "구분", "사번", "소   속", "직종", "직책", "성   명",
    "생년월일", "", "입  사  일",
    "현     장     명", "모 델 명", "모 델 번 호", "인수일자", "비 고",
]
_MERGE_SINGLE = ["A", "B", "C", "D", "E", "F", "I", "J", "K", "L", "M", "N"]


def build_data_sheet(ws, rows: list, title: str):
    ws.title = title
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    for ci, h in enumerate(_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _HFONT; c.fill = _HFILL; c.alignment = _CTR; c.border = _BRD
        ws.column_dimensions[get_column_letter(ci)].width = _WIDTHS[ci - 1]

    for ci in range(1, 15):
        c = ws.cell(row=2, column=ci)
        c.fill = _HFILL; c.border = _BRD

    # 헤더 병합: G1:H2 (생년월일), 나머지 단일 열 1행~2행
    ws.merge_cells("G1:H2")
    for col in _MERGE_SINGLE:
        ws.merge_cells(f"{col}1:{col}2")

    for ri, row in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 16
        vals = [
            ri - 2,
            row.get("사번", ""),
            row.get("소속", ""),
            row.get("직종", ""),
            row.get("직책", ""),
            row.get("성명", ""),
            row.get("생년월일", ""),
            "",                          # H열: 뒷자리 (미기재)
            row.get("입사일", ""),
            row.get("현장명", ""),
            row.get("모델명", ""),
            row.get("모델번호", ""),
            row.get("인수일자", ""),
            row.get("비고", ""),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = _DFONT; c.border = _BRD
            c.alignment = _CTR if ci in (1, 2, 4, 5, 7, 8, 9, 13) else _LFT

    ws.freeze_panes = "A3"


def build_overlap_sheet(ws, rows: list):
    ws.title = "요확인_중복"
    headers = ["순번", "성명", "사번", "소속", "직책", "비고"]
    widths  = [6, 10, 7, 14, 10, 25]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _HFONT; c.fill = _HFILL; c.alignment = _CTR; c.border = _BRD
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, info in enumerate(rows, 2):
        for ci, v in enumerate(
            [ri - 1, info["성명"], info["사번"], info["소속"], info["직책"],
             "선종합·퇴사자 양쪽 존재"], 1
        ):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = _DFONT; c.border = _BRD; c.alignment = _CTR
    ws.freeze_panes = "A2"

# ─────────────────────────────────────────────────────────────────────────────
# 메인 로직
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="선엔지니어링 노트북 지급현황 최신화")
    parser.add_argument("--hr",        required=True,             help="인사기록부 xlsx 경로")
    parser.add_argument("--new-equip", default=None, nargs="+",   help="신규 전산기기 지급현황 xlsx 경로 (여러 파일 공백으로 구분)")
    parser.add_argument("--master",    default=DEFAULT_MASTER,    help="기존 마스터 파일 경로")
    parser.add_argument("--out",       default=None,              help="출력 파일 경로 (생략 시 마스터 덮어씀)")
    args = parser.parse_args()

    out_path = args.out or args.master

    # ── 1. 인사기록부 ─────────────────────────────────────────────────────
    print(f"[1/5] 인사기록부 파싱: {args.hr}")
    active, resigned = parse_hr_file(args.hr)
    print(f"      재직자 {len(active)}명 / 퇴사자 {len(resigned)}명")

    # ── 2. 신규 지급현황 ──────────────────────────────────────────────────
    new_assignments: dict = {}  # norm_name → best assignment dict
    if args.new_equip:
        total_equip = 0
        for equip_path in args.new_equip:
            print(f"[2/5] 신규 지급현황 파싱: {equip_path}")
            equip_list = parse_equip_file(equip_path)
            print(f"      {len(equip_list)}건 인식")
            total_equip += len(equip_list)
            for a in equip_list:
                n = a["성명"]
                prev = new_assignments.get(n)
                # 날짜가 더 최근인 이력을 채택 (같으면 기존 유지)
                if prev is None or a["date_key"] > prev.get("date_key", ""):
                    new_assignments[n] = a
        print(f"      [합계] {total_equip}건, 고유 인원 {len(new_assignments)}명")
    else:
        print("[2/5] 신규 지급현황 없음 → HR 정보만 갱신, 노트북 이력은 마스터 유지")

    # ── 3. 기존 마스터 ────────────────────────────────────────────────────
    print(f"[3/5] 기존 마스터 로드: {args.master}")
    master = load_master(args.master)
    print(f"      기존 등록 {len(master)}명")

    # ── 4. 병합 ───────────────────────────────────────────────────────────
    print("[4/5] 데이터 병합 중...")
    # 선종합·퇴사자 양쪽에 이름이 있는 인원
    # → 선종합(현재 HR)을 우선: 재직자로 처리하고 참고용으로만 요확인 시트에 기록
    overlap_set = {n for n in active if n in resigned}

    changes = {
        "신규_노트북지급":      [],
        "노트북_교체변경":      [],
        "미지급→노트북이동":    [],
        "신규_미지급입사":      [],
        "HR정보_변경":          [],
        "HR에서_제외됨":        [],
        "중복_재직처리":        list(overlap_set),   # 참고용 기록
    }

    rows_laptop: list = []
    rows_none:   list = []
    rows_overlap: list = list(active[n] for n in overlap_set)   # 참고용 시트용

    for name, hr_info in active.items():
        existing = master.get(name, {})
        is_overlap = name in overlap_set  # 참고용 (처리 제외 아님)

        # ── 노트북 정보 결정: 신규 지급현황 > 기존 마스터 ──────────────
        if name in new_assignments:
            na = new_assignments[name]
            laptop = {
                "현장명":  na["현장명"],
                "모델명":  na["모델명"],
                "모델번호": na["모델번호"],
                "인수일자": na["인수일자"],
                "비고":    na["비고"],
            }
            if not existing:
                changes["신규_노트북지급"].append(name)
            elif existing.get("in_sheet") == "미지급":
                changes["미지급→노트북이동"].append(name)
            elif existing.get("모델번호") != na["모델번호"]:
                changes["노트북_교체변경"].append(name)

        elif existing.get("현장명") or existing.get("모델명"):
            # 기존 마스터 이력 유지
            laptop = {
                "현장명":  existing.get("현장명", ""),
                "모델명":  existing.get("모델명", ""),
                "모델번호": existing.get("모델번호", ""),
                "인수일자": existing.get("인수일자", ""),
                "비고":    existing.get("비고", ""),
            }
        else:
            laptop = {"현장명": "", "모델명": "", "모델번호": "", "인수일자": "", "비고": ""}
            if not existing:
                changes["신규_미지급입사"].append(name)

        # HR 정보 변경 감지
        if existing and (
            existing.get("소속", "") != hr_info["소속"] or
            existing.get("직책", "") != hr_info["직책"]
        ):
            changes["HR정보_변경"].append(name)

        row = {**hr_info, **laptop}
        has_laptop = bool(row.get("현장명") or row.get("모델명"))
        (rows_laptop if has_laptop else rows_none).append(row)

    # 마스터에 있었지만 새 HR에 없는 사람 (퇴사 추정)
    for name in master:
        if name not in active and name not in resigned:
            changes["HR에서_제외됨"].append(name)

    # 인수일자 오름차순 정렬 (빈 값은 맨 뒤)
    rows_laptop.sort(
        key=lambda x: (
            parse_date_key(x.get("인수일자", "")) == "",
            parse_date_key(x.get("인수일자", "")),
        )
    )

    # ── 5. 저장 ───────────────────────────────────────────────────────────
    print(f"[5/5] 저장: {out_path}")
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_data_sheet(wb.active, rows_laptop, "노트북보유")
    build_data_sheet(wb.create_sheet(), rows_none, "미지급")
    build_overlap_sheet(wb.create_sheet(), rows_overlap)
    wb.save(out_path)

    # 변경 로그 저장
    log_path = str(Path(out_path).parent / "update_log.json")
    result = {
        "updated_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        "master_path": str(out_path),
        "hr_file":     str(args.hr),
        "equip_files": args.new_equip if args.new_equip else [],
        "summary": {
            "재직자_총":   len(active),
            "노트북보유":  len(rows_laptop),
            "미지급":      len(rows_none),
            "요확인_중복": len(rows_overlap),
        },
        "changes": {k: v for k, v in changes.items() if v},
    }
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    # 결과를 stdout에 JSON으로 출력 (Claude가 읽어 사용자에게 보고)
    print("\n" + "=" * 60)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


if __name__ == "__main__":
    main()
